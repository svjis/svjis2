from . import utils, forms, models
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import permission_required
from django.contrib.auth.models import User, Group, Permission
from django.contrib.auth.hashers import make_password
from django.contrib import messages
from django.http import HttpResponse
from django.utils.translation import gettext_lazy as _
from django.utils.translation import gettext as gt
from django.urls import reverse
from django.views.decorators.http import require_GET, require_POST
from openpyxl import Workbook


def get_side_menu(active_item, user):
    result = []
    if user.has_perm('articles.svjis_edit_admin_company'):
        result.append(
            {
                'description': _("Company"),
                'link': reverse(admin_company_edit_view),
                'active': True if active_item == 'company' else False,
            }
        )
    if user.has_perm('articles.svjis_edit_admin_company'):
        result.append(
            {
                'description': _("Board") + f' ({models.Board.objects.count()})',
                'link': reverse(admin_board_view),
                'active': True if active_item == 'board' else False,
            }
        )
    if user.has_perm('articles.svjis_edit_admin_building'):
        result.append(
            {
                'description': _("Building"),
                'link': reverse(admin_building_edit_view),
                'active': True if active_item == 'building' else False,
            }
        )
    if user.has_perm('articles.svjis_edit_admin_building'):
        result.append(
            {
                'description': _("Entrances") + f' ({models.BuildingEntrance.objects.count()})',
                'link': reverse(admin_entrance_view),
                'active': True if active_item == 'entrances' else False,
            }
        )
    if user.has_perm('articles.svjis_edit_admin_building'):
        result.append(
            {
                'description': _("Building units") + f' ({models.BuildingUnit.objects.count()})',
                'link': reverse(admin_building_unit_view),
                'active': True if active_item == 'units' else False,
            }
        )
    if user.has_perm('articles.svjis_edit_admin_users'):
        result.append(
            {
                'description': _("Users") + f' ({User.objects.filter(is_active=True).count()})',
                'link': reverse(admin_user_view),
                'active': True if active_item == 'users' else False,
            }
        )
    if user.has_perm('articles.svjis_edit_admin_groups'):
        result.append(
            {
                'description': _("Groups"),
                'link': reverse(admin_group_view),
                'active': True if active_item == 'groups' else False,
            }
        )
    if user.has_perm('articles.svjis_edit_admin_preferences'):
        result.append(
            {
                'description': _("Preferences"),
                'link': reverse(admin_preferences_view),
                'active': True if active_item == 'preferences' else False,
            }
        )
    if user.has_perm('articles.svjis_view_admin_menu'):
        result.append(
            {
                'description': _("Waiting messages") + f' ({models.MessageQueue.objects.filter(status=0).count()})',
                'link': reverse(admin_messages_view),
                'active': True if active_item == 'messages' else False,
            }
        )
    return result


# Administration - Company
@permission_required("articles.svjis_edit_admin_company")
@require_GET
def admin_company_edit_view(request):
    instance, _created = models.Company.objects.get_or_create(pk=1)
    form = forms.CompanyForm(instance=instance)
    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Administration")
    ctx['form'] = form
    ctx['aside_menu_items'] = get_side_menu('company', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('admin', request.user)
    return render(request, "admin_company_edit.html", ctx)


@permission_required("articles.svjis_edit_admin_company")
@require_POST
def admin_company_save_view(request):
    instance, _created = models.Company.objects.get_or_create(pk=1)
    form = forms.CompanyForm(request.POST, request.FILES, instance=instance)
    if form.is_valid:
        form.save()
        messages.info(request, _('Saved'))
    else:
        for error in form.errors:
            messages.error(request, f"{_('Form validation error')}: {error}")
    return redirect(admin_company_edit_view)


# Administration - Board


@permission_required("articles.svjis_edit_admin_company")
@require_GET
def admin_board_view(request):
    board_list = models.Board.objects.select_related('member')
    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Administration")
    ctx['aside_menu_items'] = get_side_menu('board', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('admin', request.user)
    ctx['object_list'] = board_list
    return render(request, "admin_board.html", ctx)


@permission_required("articles.svjis_edit_admin_company")
@require_GET
def admin_board_edit_view(request, pk):
    if pk != 0:
        i = get_object_or_404(models.Board, pk=pk)
        form = forms.BoardForm(instance=i)
    else:
        form = forms.BoardForm

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Administration")
    ctx['form'] = form
    ctx['pk'] = pk
    ctx['aside_menu_items'] = get_side_menu('board', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('admin', request.user)
    return render(request, "admin_board_edit.html", ctx)


@permission_required("articles.svjis_edit_admin_company")
@require_POST
def admin_board_save_view(request):
    pk = int(request.POST['pk'])
    if pk == 0:
        form = forms.BoardForm(request.POST)
    else:
        instance = get_object_or_404(models.Board, pk=pk)
        form = forms.BoardForm(request.POST, instance=instance)
    if form.is_valid:
        obj = form.save(commit=False)
        obj.company = models.Company.objects.get(pk=1)
        obj.save()
    else:
        for error in form.errors:
            messages.error(request, f"{_('Form validation error')}: {error}")
    return redirect(admin_board_view)


@permission_required("articles.svjis_edit_admin_company")
@require_GET
def admin_board_delete_view(request, pk):
    obj = get_object_or_404(models.Board, pk=pk)
    obj.delete()
    return redirect(admin_board_view)


# Administration - Building
@permission_required("articles.svjis_edit_admin_building")
@require_GET
def admin_building_edit_view(request):
    instance, _created = models.Building.objects.get_or_create(pk=1)
    form = forms.BuildingForm(instance=instance)
    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Administration")
    ctx['form'] = form
    ctx['aside_menu_items'] = get_side_menu('building', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('admin', request.user)
    return render(request, "admin_building_edit.html", ctx)


@permission_required("articles.svjis_edit_admin_building")
@require_POST
def admin_building_save_view(request):
    instance, _created = models.Building.objects.get_or_create(pk=1)
    form = forms.BuildingForm(request.POST, instance=instance)
    if form.is_valid:
        form.save()
        messages.info(request, _('Saved'))
    else:
        for error in form.errors:
            messages.error(request, f"{_('Form validation error')}: {error}")
    return redirect(admin_building_edit_view)


# Administration - BuildingEntrance
@permission_required("articles.svjis_edit_admin_building")
@require_GET
def admin_entrance_view(request):
    entrance_list = models.BuildingEntrance.objects.all()
    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Administration")
    ctx['aside_menu_items'] = get_side_menu('entrances', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('admin', request.user)
    ctx['object_list'] = entrance_list
    return render(request, "admin_entrance.html", ctx)


@permission_required("articles.svjis_edit_admin_building")
@require_GET
def admin_entrance_edit_view(request, pk):
    if pk != 0:
        i = get_object_or_404(models.BuildingEntrance, pk=pk)
        form = forms.BuildingEntranceForm(instance=i)
    else:
        form = forms.BuildingEntranceForm

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Administration")
    ctx['form'] = form
    ctx['pk'] = pk
    ctx['aside_menu_items'] = get_side_menu('entrances', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('admin', request.user)
    return render(request, "admin_entrance_edit.html", ctx)


@permission_required("articles.svjis_edit_admin_building")
@require_POST
def admin_entrance_save_view(request):
    pk = int(request.POST['pk'])
    if pk == 0:
        form = forms.BuildingEntranceForm(request.POST)
    else:
        instance = get_object_or_404(models.BuildingEntrance, pk=pk)
        form = forms.BuildingEntranceForm(request.POST, instance=instance)
    if form.is_valid:
        obj = form.save(commit=False)
        obj.building, _created = models.Building.objects.get_or_create(pk=1)
        obj.save()
    else:
        for error in form.errors:
            messages.error(request, f"{_('Form validation error')}: {error}")

    return redirect(admin_entrance_view)


@permission_required("articles.svjis_edit_admin_building")
@require_GET
def admin_entrance_delete_view(request, pk):
    obj = get_object_or_404(models.BuildingEntrance, pk=pk)
    obj.delete()
    return redirect(admin_entrance_view)


# Administration - BuildingUnit
@permission_required("articles.svjis_edit_admin_building")
@require_GET
def admin_building_unit_view(request):
    unit_list = models.BuildingUnit.objects.select_related('type', 'entrance')
    type_list = models.BuildingUnitType.objects.all()
    entrance_list = models.BuildingEntrance.objects.all()

    type_filter = int(request.GET.get('type_filter', 0))
    if type_filter != 0:
        unit_list = unit_list.filter(type_id=type_filter)

    entrance_filter = int(request.GET.get('entrance_filter', 0))
    if entrance_filter != 0:
        unit_list = unit_list.filter(entrance_id=entrance_filter)

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Administration")
    ctx['aside_menu_items'] = get_side_menu('units', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('admin', request.user)
    ctx['object_list'] = unit_list.order_by('id')
    ctx['type_list'] = type_list
    ctx['type_filter'] = type_filter
    ctx['entrance_list'] = entrance_list
    ctx['entrance_filter'] = entrance_filter
    return render(request, "admin_building_unit.html", ctx)


@permission_required("articles.svjis_edit_admin_building")
@require_GET
def admin_building_unit_edit_view(request, pk):
    if pk != 0:
        i = get_object_or_404(models.BuildingUnit, pk=pk)
        form = forms.BuildingUnitForm(instance=i)
    else:
        form = forms.BuildingUnitForm

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Administration")
    ctx['form'] = form
    ctx['pk'] = pk
    ctx['aside_menu_items'] = get_side_menu('units', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('admin', request.user)
    return render(request, "admin_building_unit_edit.html", ctx)


@permission_required("articles.svjis_edit_admin_building")
@require_POST
def admin_building_unit_save_view(request):
    pk = int(request.POST['pk'])
    if pk == 0:
        form = forms.BuildingUnitForm(request.POST)
    else:
        instance = get_object_or_404(models.BuildingUnit, pk=pk)
        form = forms.BuildingUnitForm(request.POST, instance=instance)
    if form.is_valid:
        obj = form.save(commit=False)
        obj.building, _created = models.Building.objects.get_or_create(pk=1)
        obj.save()
    else:
        for error in form.errors:
            messages.error(request, f"{_('Form validation error')}: {error}")

    return redirect(admin_building_unit_view)


@permission_required("articles.svjis_edit_admin_building")
@require_GET
def admin_building_unit_delete_view(request, pk):
    obj = get_object_or_404(models.BuildingUnit, pk=pk)
    obj.delete()
    return redirect(admin_building_unit_view)


@permission_required("articles.svjis_edit_admin_building")
@require_GET
def admin_building_unit_owners_view(request, pk):
    bu = get_object_or_404(models.BuildingUnit, pk=pk)
    user_list = [u for u in User.objects.filter(is_active=True).order_by('last_name') if u not in bu.owners.all()]

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Administration")
    ctx['bu'] = bu
    ctx['pk'] = pk
    ctx['user_list'] = user_list
    ctx['aside_menu_items'] = get_side_menu('units', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('admin', request.user)
    return render(request, "admin_building_unit_owners_edit.html", ctx)


@permission_required("articles.svjis_edit_admin_building")
@require_POST
def admin_building_unit_owners_save_view(request):
    pk = int(request.POST['pk'])
    owner_id = int(request.POST['owner_id'])

    if owner_id > 0:
        bu = get_object_or_404(models.BuildingUnit, pk=pk)
        u = get_object_or_404(User, pk=owner_id)
        bu.owners.add(u)

    return redirect(admin_building_unit_owners_view, pk=pk)


@permission_required("articles.svjis_edit_admin_building")
@require_GET
def admin_building_unit_owners_delete_view(request, pk, owner):
    bu = get_object_or_404(models.BuildingUnit, pk=pk)
    u = get_object_or_404(User, pk=owner)
    bu.owners.remove(u)
    return redirect(admin_building_unit_owners_view, pk=pk)


@permission_required("articles.svjis_edit_admin_building")
@require_GET
def admin_building_unit_export_to_excel_view(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Building_Units.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = gt("Building units")

    # Add headers
    headers = [
        gt("Type"),
        gt("Entrance"),
        gt("Registration Id"),
        gt("Description"),
        gt("Numerator"),
        gt("Denominator"),
    ]
    ws.append(headers)

    header_st = utils.get_worksheet_header_style()

    for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
        for cell in rows:
            cell.style = header_st

    # Add data from the model
    unit_list = models.BuildingUnit.objects.all().order_by('id')
    for u in unit_list:
        unit_entrance = u.entrance.description if u.entrance else ''
        ws.append([u.type.description, unit_entrance, u.registration_id, u.description, u.numerator, u.denominator])

    utils.adjust_worksheet_columns_width(ws)

    # Save the workbook to the HttpResponse
    wb.save(response)
    return response


# Administration - User
@permission_required("articles.svjis_edit_admin_users")
@require_GET
def admin_user_view(request):
    deactivated_users = request.GET.get('deactivated_users', False) == 'on'
    user_list = User.objects.filter(is_active=not deactivated_users)
    group_filter = int(request.GET.get('group_filter', 0))
    if group_filter != 0:
        g = Group.objects.filter(pk=group_filter)
        user_list = User.objects.filter(groups__in=g)
    group_list = Group.objects.all()

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Administration")
    ctx['aside_menu_items'] = get_side_menu('users', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('admin', request.user)
    ctx['object_list'] = user_list.order_by('last_name', 'first_name')
    ctx['group_list'] = group_list.order_by('name')
    ctx['group_filter'] = group_filter
    ctx['deactivated_users'] = deactivated_users
    return render(request, "admin_user.html", ctx)


@permission_required("articles.svjis_edit_admin_users")
@require_GET
def admin_user_edit_view(request, pk):
    if pk != 0:
        instance = get_object_or_404(User, pk=pk)
        uform = forms.UserForm(instance=instance)
        pinstance, _created = models.UserProfile.objects.get_or_create(user=instance)
        pform = forms.UserProfileForm(instance=pinstance)
    else:
        instance = User
        uform = forms.UserForm()
        pform = forms.UserProfileForm()

    group_list = []
    user_group_list = []
    if pk != 0:
        user_group_list = Group.objects.filter(user__id=instance.id)
    for g in Group.objects.all().order_by('name'):
        item = {'name': g.name, 'checked': g in user_group_list}
        group_list.append(item)

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Administration")
    ctx['uform'] = uform
    ctx['pform'] = pform
    ctx['group_list'] = group_list
    ctx['pk'] = pk
    ctx['aside_menu_items'] = get_side_menu('users', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('admin', request.user)
    return render(request, "admin_user_edit.html", ctx)


@permission_required("articles.svjis_edit_admin_users")
@require_POST
def admin_user_save_view(request):
    pk = int(request.POST['pk'])
    if pk != 0:
        user_i = get_object_or_404(User, pk=pk)
        user_form = forms.UserForm(request.POST, instance=user_i)
        user_profile_form = forms.UserProfileForm(request.POST, instance=user_i.userprofile)
    else:
        user_form = forms.UserForm(request.POST)
        user_profile_form = forms.UserProfileForm(request.POST)

    if user_form.is_valid() and user_profile_form.is_valid():
        u = user_form.save()
        u.userprofile = user_profile_form.instance
        u.userprofile.save()

        password = request.POST.get('password', '')
        if password != '':
            u.password = make_password(password)
            u.save()

        send_credentials = request.POST.get('send_credentials', False) == 'on'
        if send_credentials:
            utils.send_new_password(u)

        # Set groups
        user_group_list = Group.objects.filter(user__id=u.id)
        for g in Group.objects.all():
            group_set = request.POST.get(g.name, False) == 'on'
            if group_set and g not in user_group_list:
                u.groups.add(g)
            if not group_set and g in user_group_list:
                u.groups.remove(g)

    else:
        for error in user_form.errors:
            messages.error(request, f"{_('Form validation error')}: {error}")
        for error in user_profile_form.errors:
            messages.error(request, f"{_('Form validation error')}: {error}")
        return redirect(reverse('admin_user_edit', kwargs={'pk': pk}))

    return redirect(admin_user_view)


@permission_required("articles.svjis_edit_admin_users")
@require_GET
def admin_user_owns_view(request, pk):
    u = get_object_or_404(User, pk=pk)
    bu_list = [bu for bu in models.BuildingUnit.objects.all().order_by('id') if bu not in u.buildingunit_set.all()]
    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Administration")
    ctx['u'] = u
    ctx['pk'] = pk
    ctx['bu_list'] = bu_list
    ctx['aside_menu_items'] = get_side_menu('users', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('admin', request.user)
    return render(request, "admin_user_ownes_edit.html", ctx)


@permission_required("articles.svjis_edit_admin_users")
@require_POST
def admin_user_owns_save_view(request):
    pk = int(request.POST['pk'])
    owner_id = int(request.POST['owner_id'])

    if owner_id > 0:
        u = get_object_or_404(User, pk=pk)
        bu = get_object_or_404(models.BuildingUnit, pk=owner_id)
        u.buildingunit_set.add(bu)

    return redirect(admin_user_owns_view, pk=pk)


@permission_required("articles.svjis_edit_admin_users")
@require_GET
def admin_user_owns_delete_view(request, pk, owner):
    u = get_object_or_404(User, pk=pk)
    bu = get_object_or_404(models.BuildingUnit, pk=owner)
    u.buildingunit_set.remove(bu)
    return redirect(admin_user_owns_view, pk=pk)


@permission_required("articles.svjis_edit_admin_users")
@require_GET
def admin_user_export_to_excel_view(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Users.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = gt("Users")

    # Add headers
    headers = [
        gt("Salutation"),
        gt("First name"),
        gt("Last name"),
        gt("Address"),
        gt("City"),
        gt("Post code"),
        gt("Country"),
        gt("Phone"),
        gt("Email address"),
        gt("Username"),
    ]
    ws.append(headers)

    header_st = utils.get_worksheet_header_style()

    for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
        for cell in rows:
            cell.style = header_st

    # Add data from the model
    user_list = User.objects.filter(is_active=True).order_by('last_name', 'first_name')
    for u in user_list:
        if hasattr(u, 'userprofile'):
            ws.append(
                [
                    u.userprofile.salutation,
                    u.first_name,
                    u.last_name,
                    u.userprofile.address,
                    u.userprofile.city,
                    u.userprofile.post_code,
                    u.userprofile.country,
                    u.userprofile.phone,
                    u.email,
                    u.username,
                ]
            )

    utils.adjust_worksheet_columns_width(ws)

    # Save the workbook to the HttpResponse
    wb.save(response)
    return response


# Administration - Group
@permission_required("articles.svjis_edit_admin_groups")
@require_GET
def admin_group_view(request):
    group_list = Group.objects.all()
    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Administration")
    ctx['aside_menu_items'] = get_side_menu('groups', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('admin', request.user)
    ctx['object_list'] = group_list.order_by('name')
    return render(request, "admin_group.html", ctx)


@permission_required("articles.svjis_edit_admin_groups")
@require_GET
def admin_group_edit_view(request, pk):
    if pk != 0:
        i = get_object_or_404(Group, pk=pk)
        form = forms.GroupEditForm(instance=i)
    else:
        i = Group
        form = forms.GroupEditForm

    permission_list = []
    group_perm_list = []
    if pk != 0:
        group_perm_list = Permission.objects.filter(group__id=i.id)
    for p in Permission.objects.all():
        if p.codename.startswith('svjis_'):
            item = {'name': p.codename, 'checked': p in group_perm_list}
            permission_list.append(item)

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Administration")
    ctx['form'] = form
    ctx['instance'] = i
    ctx['permission_list'] = permission_list
    ctx['pk'] = pk
    ctx['aside_menu_items'] = get_side_menu('groups', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('admin', request.user)
    return render(request, "admin_group_edit.html", ctx)


@permission_required("articles.svjis_edit_admin_groups")
@require_POST
def admin_group_save_view(request):
    pk = int(request.POST['pk'])
    if pk == 0:
        form = forms.GroupEditForm(request.POST)
    else:
        instance = get_object_or_404(Group, pk=pk)
        form = forms.GroupEditForm(request.POST, instance=instance)
    if form.is_valid:
        instance = form.save()

        # Set permissions
        group_perm_list = Permission.objects.filter(group__id=instance.id)
        for p in Permission.objects.all():
            if p.codename.startswith('svjis_'):
                perm_set = request.POST.get(p.codename, False) == 'on'
                if perm_set and p not in group_perm_list:
                    instance.permissions.add(p)
                if not perm_set and p in group_perm_list:
                    instance.permissions.remove(p)
    else:
        for error in form.errors:
            messages.error(request, f"{_('Form validation error')}: {error}")

    return redirect(admin_group_view)


@permission_required("articles.svjis_edit_admin_groups")
@require_GET
def admin_group_delete_view(request, pk):
    obj = get_object_or_404(Group, pk=pk)
    obj.delete()
    return redirect(admin_group_view)


# Administration - Preferences
@permission_required("articles.svjis_edit_admin_preferences")
@require_GET
def admin_preferences_view(request):
    property_list = models.Preferences.objects.all()
    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Administration")
    ctx['aside_menu_items'] = get_side_menu('preferences', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('admin', request.user)
    ctx['object_list'] = property_list.order_by('key')
    return render(request, "admin_preferences.html", ctx)


@permission_required("articles.svjis_edit_admin_preferences")
@require_GET
def admin_preferences_edit_view(request, pk):
    if pk != 0:
        i = get_object_or_404(models.Preferences, pk=pk)
        form = forms.PreferencesForm(instance=i)
    else:
        form = forms.PreferencesForm

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Administration")
    ctx['form'] = form
    ctx['pk'] = pk
    ctx['aside_menu_items'] = get_side_menu('preferences', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('admin', request.user)
    return render(request, "admin_preferences_edit.html", ctx)


@permission_required("articles.svjis_edit_admin_preferences")
@require_POST
def admin_preferences_save_view(request):
    pk = int(request.POST['pk'])
    if pk == 0:
        form = forms.PreferencesForm(request.POST)
    else:
        instance = get_object_or_404(models.Preferences, pk=pk)
        form = forms.PreferencesForm(request.POST, instance=instance)
    if form.is_valid:
        form.save()
    else:
        for error in form.errors:
            messages.error(request, f"{_('Form validation error')}: {error}")
    return redirect(admin_preferences_view)


@permission_required("articles.svjis_edit_admin_preferences")
@require_GET
def admin_preferences_delete_view(request, pk):
    obj = get_object_or_404(models.Preferences, pk=pk)
    obj.delete()
    return redirect(admin_preferences_view)


# Administration - Waiting messages
@permission_required("articles.svjis_view_admin_menu")
@require_GET
def admin_messages_view(request):
    message_list = models.MessageQueue.objects.filter(status=0)
    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Administration")
    ctx['aside_menu_items'] = get_side_menu('messages', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('admin', request.user)
    ctx['object_list'] = message_list.order_by('pk')
    return render(request, "admin_messages.html", ctx)

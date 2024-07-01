from . import utils, models, forms
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import permission_required
from django.contrib.auth.models import Group, User
from django.core.paginator import Paginator, InvalidPage
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.utils import timezone, dateformat
from django.utils.translation import gettext_lazy as _
from django.utils.translation import gettext as gt
from django.views.decorators.http import require_GET, require_POST
from openpyxl import Workbook
from openpyxl.styles import Font


def get_side_menu(active_item, user):
    result = []
    if user.has_perm('articles.svjis_edit_article'):
        result.append(
            {
                'description': _("Articles"),
                'link': reverse(redaction_article_view),
                'active': True if active_item == 'article' else False,
            }
        )
    if user.has_perm('articles.svjis_edit_article_news'):
        result.append(
            {
                'description': _("News"),
                'link': reverse(redaction_news_view),
                'active': True if active_item == 'news' else False,
            }
        )
    if user.has_perm('articles.svjis_edit_useful_link'):
        result.append(
            {
                'description': _("Useful Links"),
                'link': reverse(redaction_useful_link_view),
                'active': True if active_item == 'links' else False,
            }
        )
    if user.has_perm('articles.svjis_edit_survey'):
        result.append(
            {
                'description': _("Surveys"),
                'link': reverse(redaction_survey_view),
                'active': True if active_item == 'surveys' else False,
            }
        )
    if user.has_perm('articles.svjis_edit_article_menu'):
        result.append(
            {
                'description': _("Menu"),
                'link': reverse(redaction_menu_view),
                'active': True if active_item == 'menu' else False,
            }
        )
    return result


# Redaction - Article Menu
def get_article_menu():
    result = []
    menu_items = models.ArticleMenu.objects.all()
    level = 0
    for i in menu_items:
        if i.parent is None:
            node = {'item': i, 'level': level, 'articles': i.articles.count()}
            result.append(node)
            result.extend(get_article_submenu(i, menu_items, level + 1))
    return result


def get_article_submenu(parent, menu_items, level):
    result = []
    for i in menu_items:
        if i.parent == parent:
            node = {'item': i, 'level': level, 'articles': i.articles.count()}
            result.append(node)
            result.extend(get_article_submenu(i, menu_items, level + 1))
    return result


@permission_required("articles.svjis_edit_article_menu")
@require_GET
def redaction_menu_view(request):
    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Redaction")
    ctx['aside_menu_items'] = get_side_menu('menu', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('redaction', request.user)
    ctx['object_list'] = get_article_menu()
    return render(request, "redaction_menu.html", ctx)


@permission_required("articles.svjis_edit_article_menu")
@require_GET
def redaction_menu_edit_view(request, pk):
    if pk != 0:
        am = get_object_or_404(models.ArticleMenu, pk=pk)
        form = forms.ArticleMenuForm(instance=am)
    else:
        form = forms.ArticleMenuForm

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Redaction")
    ctx['form'] = form
    ctx['pk'] = pk
    ctx['aside_menu_items'] = get_side_menu('menu', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('redaction', request.user)
    return render(request, "redaction_menu_edit.html", ctx)


@permission_required("articles.svjis_edit_article_menu")
@require_POST
def redaction_menu_save_view(request):
    pk = int(request.POST['pk'])
    if pk == 0:
        form = forms.ArticleMenuForm(request.POST)
    else:
        instance = get_object_or_404(models.ArticleMenu, pk=pk)
        form = forms.ArticleMenuForm(request.POST, instance=instance)

    if form.is_valid():
        form.save()
    else:
        for error in form.errors:
            messages.error(request, f"{_('Form validation error')}: {error}")

    return redirect(redaction_menu_view)


@permission_required("articles.svjis_edit_article_menu")
@require_GET
def redaction_menu_delete_view(request, pk):
    obj = get_object_or_404(models.ArticleMenu, pk=pk)
    obj.delete()
    return redirect(redaction_menu_view)


# Redaction - Article
@permission_required("articles.svjis_edit_article")
@require_GET
def redaction_article_view(request):
    article_list = models.Article.objects.all()

    # Search
    search = request.GET.get('search')
    if search is not None and len(search) < 3:
        messages.error(request, _("Search: Keyword '{}' is too short. Type at least 3 characters.").format(search))
        search = None
    if search is not None and len(search) > 100:
        messages.error(request, _("Search: Keyword is too long. Type maximum of 100 characters."))
        search = None
    if search is not None:
        article_list = article_list.filter(
            Q(header__icontains=search) | Q(perex__icontains=search) | Q(body__icontains=search)
        )
        header = _("Search results") + f": {search}"
    else:
        search = ''
        header = _("Article")

    # Paginator
    is_paginated = len(article_list) > getattr(settings, 'SVJIS_ARTICLE_PAGE_SIZE', 10)
    page = request.GET.get('page', 1)
    paginator = Paginator(article_list, per_page=getattr(settings, 'SVJIS_ARTICLE_PAGE_SIZE', 10))
    page_obj = paginator.get_page(page)
    try:
        article_list = paginator.page(page)
    except InvalidPage:
        article_list = paginator.page(paginator.num_pages)
    page_parameter = '' if search == '' else f"search={search}"

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Redaction")
    ctx['is_paginated'] = is_paginated
    ctx['page_obj'] = page_obj
    ctx['page_parameter'] = page_parameter
    ctx['search_endpoint'] = reverse(redaction_article_view)
    ctx['search'] = search
    ctx['header'] = header
    ctx['aside_menu_items'] = get_side_menu('article', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('redaction', request.user)
    ctx['object_list'] = article_list
    return render(request, "redaction_article.html", ctx)


@permission_required("articles.svjis_edit_article")
@require_GET
def redaction_article_edit_view(request, pk):
    if pk != 0:
        a = get_object_or_404(models.Article, pk=pk)
        form = forms.ArticleForm(instance=a)
    else:
        form = forms.ArticleForm

    group_list = []
    for g in Group.objects.all().order_by('name'):
        item = {'name': g.name, 'checked': g in form.instance.visible_for_group.all() if pk != 0 else False}
        group_list.append(item)

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Redaction")
    ctx['form'] = form
    ctx['group_list'] = group_list
    ctx['asset_form'] = forms.ArticleAssetForm
    ctx['pk'] = pk
    if pk != 0:
        ctx['assets'] = utils.wrap_assets(form.instance.assets)
    ctx['aside_menu_items'] = get_side_menu('article', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('redaction', request.user)
    return render(request, "redaction_article_edit.html", ctx)


@permission_required("articles.svjis_edit_article")
@require_POST
def redaction_article_save_view(request):
    pk = int(request.POST['pk'])
    if pk == 0:
        form = forms.ArticleForm(request.POST)
    else:
        instance = get_object_or_404(models.Article, pk=pk)
        form = forms.ArticleForm(request.POST, instance=instance)

    if form.is_valid():
        obj = form.save(commit=False)
        if pk == 0:
            obj.author = request.user
        obj.save()

        # Set groups
        group_list = obj.visible_for_group.all()
        for g in Group.objects.all():
            gr_set = request.POST.get(g.name, False) == 'on'
            if gr_set and g not in group_list:
                obj.visible_for_group.add(g)
            if not gr_set and g in group_list:
                obj.visible_for_group.remove(g)

        # Set watching users
        if pk == 0:
            obj.watching_users.add(request.user)
    else:
        for error in form.errors:
            messages.error(request, f"{_('Form validation error')}: {error}")
        return redirect(reverse('redaction_article_edit', kwargs={'pk': pk}))

    return redirect(redaction_article_view)


def get_users_for_notification(article):
    users = []
    if article.published:
        users = User.objects.filter(is_active=True).exclude(email='').distinct().order_by('last_name')
        if not article.visible_for_all:
            groups = article.visible_for_group.all()
            q = Q(groups__in=groups)
            users = users.filter(q)
    return users


@permission_required("articles.svjis_edit_article")
@require_GET
def redaction_article_notifications_view(request, pk):
    article = get_object_or_404(models.Article, pk=pk)
    users = get_users_for_notification(article)

    ctx = utils.get_context()
    ctx['article'] = article
    ctx['object_list'] = users
    ctx['aside_menu_name'] = _("Redaction")
    ctx['aside_menu_items'] = get_side_menu('article', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('redaction', request.user)
    return render(request, "redaction_article_notifications.html", ctx)


@permission_required("articles.svjis_edit_article")
@require_POST
def redaction_article_notifications_send_view(request):
    pk = int(request.POST['pk'])
    article = get_object_or_404(models.Article, pk=pk)
    users = get_users_for_notification(article)

    recipients = [u for u in users if request.POST.get(f"u_{u.pk}", False) == 'on']
    utils.send_article_notification(recipients, f"{request.scheme}://{request.get_host()}", article)
    i = len(recipients)

    ctx = utils.get_context()
    ctx['article'] = article
    ctx['num_of_recipients'] = i
    ctx['aside_menu_name'] = _("Redaction")
    ctx['aside_menu_items'] = get_side_menu('article', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('redaction', request.user)
    return render(request, "redaction_article_notifications_sent.html", ctx)


# Redaction - ArticleAsset
@permission_required("articles.svjis_edit_article")
@require_POST
def redaction_article_asset_save_view(request):
    article_pk = int(request.POST.get('article_pk'))
    article = get_object_or_404(models.Article, pk=article_pk)
    form = forms.ArticleAssetForm(request.POST, request.FILES)
    if form.is_valid():
        obj = form.save(commit=False)
        obj.article = article
        obj.save()
    else:
        for error in form.errors:
            messages.error(request, error)
    return redirect(redaction_article_edit_view, pk=article_pk)


@permission_required("articles.svjis_edit_article")
@require_GET
def redaction_article_asset_delete_view(request, pk):
    obj = get_object_or_404(models.ArticleAsset, pk=pk)
    article_pk = obj.article.pk
    obj.delete()
    return redirect(redaction_article_edit_view, pk=article_pk)


# Redaction - MiniNews
@permission_required("articles.svjis_edit_article_news")
@require_GET
def redaction_news_view(request):
    news_list = models.News.objects.all()
    header = _("News")

    # Paginator
    is_paginated = len(news_list) > getattr(settings, 'SVJIS_NEWS_PAGE_SIZE', 10)
    page = request.GET.get('page', 1)
    paginator = Paginator(news_list, per_page=getattr(settings, 'SVJIS_NEWS_PAGE_SIZE', 10))
    page_obj = paginator.get_page(page)
    try:
        news_list = paginator.page(page)
    except InvalidPage:
        news_list = paginator.page(paginator.num_pages)

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Redaction")
    ctx['is_paginated'] = is_paginated
    ctx['page_obj'] = page_obj
    ctx['header'] = header
    ctx['aside_menu_items'] = get_side_menu('news', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('redaction', request.user)
    ctx['object_list'] = news_list
    return render(request, "redaction_news.html", ctx)


@permission_required("articles.svjis_edit_article_news")
@require_GET
def redaction_news_edit_view(request, pk):
    if pk != 0:
        a = get_object_or_404(models.News, pk=pk)
        form = forms.NewsForm(instance=a)
    else:
        form = forms.NewsForm

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Redaction")
    ctx['form'] = form
    ctx['pk'] = pk
    ctx['aside_menu_items'] = get_side_menu('news', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('redaction', request.user)
    return render(request, "redaction_news_edit.html", ctx)


@permission_required("articles.svjis_edit_article_news")
@require_POST
def redaction_news_save_view(request):
    pk = int(request.POST['pk'])
    if pk == 0:
        form = forms.NewsForm(request.POST)
    else:
        instance = get_object_or_404(models.News, pk=pk)
        form = forms.NewsForm(request.POST, instance=instance)

    if form.is_valid():
        obj = form.save(commit=False)
        if pk == 0:
            obj.author = request.user
        obj.save()
    else:
        for error in form.errors:
            messages.error(request, error)
    return redirect(redaction_news_view)


@permission_required("articles.svjis_edit_article_news")
@require_GET
def redaction_news_delete_view(request, pk):
    obj = get_object_or_404(models.News, pk=pk)
    obj.delete()
    return redirect(redaction_news_view)


# Redaction - UsefulLink
@permission_required("articles.svjis_edit_useful_link")
@require_GET
def redaction_useful_link_view(request):
    useful_link_list = models.UsefulLink.objects.all()
    header = _("Useful links")

    # Paginator
    is_paginated = len(useful_link_list) > getattr(settings, 'SVJIS_USEFUL_LIST_PAGE_SIZE', 10)
    page = request.GET.get('page', 1)
    paginator = Paginator(useful_link_list, per_page=getattr(settings, 'SVJIS_USEFUL_LIST_PAGE_SIZE', 10))
    page_obj = paginator.get_page(page)
    try:
        useful_link_list = paginator.page(page)
    except InvalidPage:
        useful_link_list = paginator.page(paginator.num_pages)

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Redaction")
    ctx['is_paginated'] = is_paginated
    ctx['page_obj'] = page_obj
    ctx['header'] = header
    ctx['aside_menu_items'] = get_side_menu('links', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('redaction', request.user)
    ctx['object_list'] = useful_link_list
    return render(request, "redaction_useful_link.html", ctx)


@permission_required("articles.svjis_edit_useful_link")
@require_GET
def redaction_useful_link_edit_view(request, pk):
    if pk != 0:
        a = get_object_or_404(models.UsefulLink, pk=pk)
        form = forms.UsefulLinkForm(instance=a)
    else:
        form = forms.UsefulLinkForm

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Redaction")
    ctx['form'] = form
    ctx['pk'] = pk
    ctx['aside_menu_items'] = get_side_menu('links', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('redaction', request.user)
    return render(request, "redaction_useful_link_edit.html", ctx)


@permission_required("articles.svjis_edit_useful_link")
@require_POST
def redaction_useful_link_save_view(request):
    pk = int(request.POST['pk'])
    if pk == 0:
        form = forms.UsefulLinkForm(request.POST)
    else:
        instance = get_object_or_404(models.UsefulLink, pk=pk)
        form = forms.UsefulLinkForm(request.POST, instance=instance)

    if form.is_valid():
        form.save()
    else:
        for error in form.errors:
            messages.error(request, error)
    return redirect(redaction_useful_link_view)


@permission_required("articles.svjis_edit_useful_link")
@require_GET
def redaction_useful_link_delete_view(request, pk):
    obj = get_object_or_404(models.UsefulLink, pk=pk)
    obj.delete()
    return redirect(redaction_useful_link_view)


# Redaction - Surveys
@permission_required("articles.svjis_edit_survey")
@require_GET
def redaction_survey_view(request):
    survey_list = models.Survey.objects.all()
    header = _("Surveys")

    # Paginator
    is_paginated = len(survey_list) > getattr(settings, 'SVJIS_SURVEY_PAGE_SIZE', 10)
    page = request.GET.get('page', 1)
    paginator = Paginator(survey_list, per_page=getattr(settings, 'SVJIS_SURVEY_PAGE_SIZE', 10))
    page_obj = paginator.get_page(page)
    try:
        survey_list = paginator.page(page)
    except InvalidPage:
        survey_list = paginator.page(paginator.num_pages)

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Redaction")
    ctx['is_paginated'] = is_paginated
    ctx['page_obj'] = page_obj
    ctx['header'] = header
    ctx['aside_menu_items'] = get_side_menu('surveys', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('redaction', request.user)
    ctx['object_list'] = survey_list
    return render(request, "redaction_survey.html", ctx)


@permission_required("articles.svjis_edit_survey")
@require_GET
def redaction_survey_edit_view(request, pk):
    if pk != 0:
        a = get_object_or_404(models.Survey, pk=pk)
        form = forms.SurveyForm(instance=a)
        n = a.options.count()
    else:
        form = forms.SurveyForm
        n = 0

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Redaction")
    ctx['form'] = form
    ctx['pk'] = pk
    ctx['new_option_no'] = n + 1
    ctx['aside_menu_items'] = get_side_menu('surveys', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('redaction', request.user)
    return render(request, "redaction_survey_edit.html", ctx)


@permission_required("articles.svjis_edit_survey")
@require_POST
def redaction_survey_save_view(request):
    pk = int(request.POST['pk'])
    if pk == 0:
        form = forms.SurveyForm(request.POST)
    else:
        instance = get_object_or_404(models.Survey, pk=pk)
        form = forms.SurveyForm(request.POST, instance=instance)

    if form.is_valid():
        obj = form.save(commit=False)
        if pk == 0:
            obj.author = request.user
        obj.save()

        # Options
        i = 1
        while f'oid_{i}' in request.POST:
            o_pk = int(request.POST[f'oid_{i}'])
            o_description = request.POST.get(f'o_{i}', '')
            if o_pk != 0:
                o_i = get_object_or_404(models.SurveyOption, pk=o_pk)
                o_i.description = o_description
                o_i.save()
            else:
                if o_description != '':
                    models.SurveyOption.objects.create(survey=obj, description=o_description)
            i += 1

    else:
        for error in form.errors:
            messages.error(request, error)
    return redirect(redaction_survey_view)


@permission_required("articles.svjis_edit_survey")
@require_GET
def redaction_survey_delete_view(request, pk):
    obj = get_object_or_404(models.Survey, pk=pk)
    obj.delete()
    return redirect(redaction_survey_view)


@permission_required("articles.svjis_edit_survey")
@require_GET
def redaction_survey_option_delete_view(request, pk):
    obj = get_object_or_404(models.SurveyOption, pk=pk)
    survey_pk = obj.survey.pk
    obj.delete()
    return redirect(redaction_survey_edit_view, pk=survey_pk)


@permission_required("articles.svjis_edit_survey")
@require_GET
def redaction_survey_results_view(request, pk):
    header = _("Survey")
    survey = get_object_or_404(models.Survey, pk=pk)
    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Redaction")
    ctx['obj'] = survey
    ctx['header'] = header
    ctx['aside_menu_items'] = get_side_menu('surveys', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('redaction', request.user)
    return render(request, "redaction_survey_results.html", ctx)


@permission_required("articles.svjis_edit_survey")
@require_GET
def redaction_survey_results_export_to_excel_view(request, pk):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Survey_Results.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = gt("Survey")

    survey = get_object_or_404(models.Survey, pk=pk)

    # Survey
    ws['A1'] = gt("Survey")
    ws['A1'].font = Font(bold=True)
    ws['A2'] = survey.description

    # Result
    ws['A4'] = gt("Result")
    ws['A4'].font = Font(bold=True)

    headers = [gt("Description"), gt("Votes"), gt("Votes") + ' %']
    ws.append(headers)

    header_st = utils.get_worksheet_header_style()
    for rows in ws.iter_rows(min_row=5, max_row=5, min_col=1, max_col=len(headers)):
        for cell in rows:
            cell.style = header_st

    for o in survey.options:
        ws.append([o.description, o.total, round(o.pct, 2)])

    # Log
    last = len(ws['A']) + 1
    ws['A' + str(last + 1)] = gt("Voting log")
    ws['A' + str(last + 1)].font = Font(bold=True)

    headers = [gt("Time"), gt("User"), gt("Option")]
    ws.append(headers)

    last = len(ws['A'])
    for rows in ws.iter_rows(min_row=last, max_row=last, min_col=1, max_col=len(headers)):
        for cell in rows:
            cell.style = header_st

    for a in survey.answers:
        time = dateformat.format(timezone.localtime(a.time), "d.m.Y H:i")
        ws.append([time, f'{a.user.first_name} {a.user.last_name}', a.option.description])

    # Save the workbook to the HttpResponse
    utils.adjust_worksheet_columns_width(ws, 50)
    wb.save(response)
    return response
